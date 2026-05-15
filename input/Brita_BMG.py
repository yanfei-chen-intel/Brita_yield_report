import pandas as pd
import plotly.express as px
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

input_file = r".\output\SOC_SCAN\SOC_SCAN_units_batch_summary.csv"
#output_file = "HRY_%s" % input_file
output_file = r".\output\SOC_SCAN\HRY_soc_scan.csv"

data = pd.read_csv(input_file)
x = data['WAFER']
y = data['Indicator']
wafer = []
partition = []
partition_PH6 = []
partition_PHRST = []
partitiion_ATPG = []
partitiion_TATPG = []

for wafer_number in x:
    if wafer_number not in wafer:
        wafer.append(wafer_number)
for partition_grp in y:
    if '::NOM::' in partition_grp:
        if partition_grp.split('::')[4] not in partition:
            partition.append(partition_grp.split('::')[4])
    if '::DDR::16G::' in partition_grp or '::DDR::18G::' in partition_grp or '::DDR::14G::' in partition_grp or '::DDR::20G::' in partition_grp or '::DDR::19G::' in partition_grp:
        if partition_grp.split('::')[3] + '::'+ partition_grp.split('::')[4] not in partition:
            partition.append(partition_grp.split('::')[3] + '::'+ partition_grp.split('::')[4])
for par in partition:
    partition_PH6.append(par + 'PH')
    partition_PHRST.append(par + 'RST')
    partitiion_ATPG.append(par + '_ATPG')
    partitiion_TATPG.append(par + '_TATPG')
table = []
with open(input_file, 'r') as f:
    read = csv.reader(f)
    for row in read:
        table.append(row)
wafer_row_ATPG = []
wafer_row_TATPG = []
for row1 in table:
    if row1[8] == 'LATEST' and row1[11] == '6' and row1[12] == 'PASS' and row1[17] != 'NaN':
        for par1 in partition:
            if (':' + par1 + ':' ) in row1[10] and '::ATPG' in row1[10]:
                row_updated = []
                for indicater in row1:
                    if indicater == row1[10]:
                        row_updated.append(par1+'_ATPG')
                    else:
                        row_updated.append(indicater)
                if row_updated not in wafer_row_ATPG:
                    wafer_row_ATPG.append(row_updated)
            elif (':' + par1 +':') in row1[10] and 'TATPG' in row1[10]:
                row_updated = []
                for indicater in row1:
                    if indicater == row1[10]:
                        row_updated.append(par1+'_TATPG')
                    else:
                        row_updated.append(indicater)
                if row_updated not in wafer_row_ATPG:
                    wafer_row_TATPG.append(row_updated)
ATPG_merge = []
TATPG_merge = []
for par2 in partitiion_ATPG:
    ATPG_NEW = [0,0,0]
    row_new = []
    row_final = []
    for row2 in wafer_row_ATPG:
        if row2[10] == par2:
            ATPG_NEW[0] += int(row2[13])
            ATPG_NEW[1] += int(row2[14])
            ATPG_NEW[2] += int(row2[16])
            row_new = row2
    for i in range(len(row_new)-3):
        if i == 1:
            row_final.append('ALL')
        elif i == 13:
            row_final.append(str(ATPG_NEW[0]))
        elif i == 14:
            row_final.append(str(ATPG_NEW[1]))
        elif i == 16:
            row_final.append(str(ATPG_NEW[2]))
        else:
            row_final.append(row_new[i])
    if ATPG_NEW[2] != 0:
        row_final.append('%.2f%%'%((ATPG_NEW[0]/ATPG_NEW[2])*100))
        ATPG_merge.append(row_final)
for par3 in partitiion_TATPG:
    TATPG_NEW = [0,0,0]
    row_new = []
    row_final = []
    for row3 in wafer_row_TATPG:
        if row3[10] == par3:
            TATPG_NEW[0] += int(row3[13])
            TATPG_NEW[1] += int(row3[14])
            TATPG_NEW[2] += int(row3[16])
            row_new = row3
    for i in range(len(row_new)-3):
        if i == 1:
            row_final.append('ALL')
        elif i == 13:
            row_final.append(str(TATPG_NEW[0]))
        elif i == 14:
            row_final.append(str(TATPG_NEW[1]))
        elif i == 16:
            row_final.append(str(TATPG_NEW[2]))
        else:
            row_final.append(row_new[i])
    if TATPG_NEW[2] != 0:
        row_final.append('%.2f%%'%((TATPG_NEW[0]/TATPG_NEW[2])*100))
        TATPG_merge.append(row_final)
Total = []
Total.append(table[0][:17])
Total[0].append('Yield')
Total[0][13] = 'Passing Unit'
Total[0][16] = 'Unit QTY'
for row4 in ATPG_merge:
    Total.append(row4)
for row5 in TATPG_merge:
    Total.append(row5)
with open(output_file, 'w', newline = "") as csvfile:
    s = csv.writer(csvfile)
    for row6 in Total:
        s.writerow(row6)


df = pd.read_csv(output_file)
df['Yield'] = df['Yield'].str.rstrip('%').astype(float)
df.sort_values('Yield', ascending=True, inplace=True)
df.to_csv(output_file, index=False)

# Export to Excel with red font for Yield < 99%
xlsx_file = r".\output\SOC_SCAN\HRY_soc_scan.xlsx"
df.to_excel(xlsx_file, index=False)
wb = load_workbook(xlsx_file)
ws = wb.active
yield_col = [cell.column for cell in ws[1] if cell.value == 'Yield'][0]

# Header: bold white font + blue background
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="0068B5")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data rows: red font for Yield < 99%
red_font = Font(color="FF0000")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[yield_col - 1].value is not None and row[yield_col - 1].value < 99:
        for cell in row:
            cell.font = red_font
wb.save(xlsx_file)

a1 = px.box(df
            , x='Indicator'
            , y='Yield'
            , points='all', title='Yield'
            )
# a1.show()
a1.write_html(r".\output\SOC_SCAN\yield_plot.html")














